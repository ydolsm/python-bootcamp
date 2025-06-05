from openpyxl import load_workbook


workbook = load_workbook("sample.xlsx")
sheet = workbook["Additional"]

sheet["A1"] = "Tickets"
print(sheet["A1"].value)

cell = sheet.cell(row=1, column=2)
cell.value = 100
print(cell.value)

workbook.save("sample.xlsx")
