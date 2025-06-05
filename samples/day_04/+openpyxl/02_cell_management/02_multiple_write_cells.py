from openpyxl import load_workbook


workbook = load_workbook("sample.xlsx")
sheet = workbook["Additional"]

tickets = {"HR": 30, "Legal": 23, "Sales": 34, "Admin": 13}

for i, (group, count) in enumerate(tickets.items(), start=3):
    sheet.cell(row=i, column=1).value = group
    sheet.cell(row=i, column=2).value = count

workbook.save("sample.xlsx")
