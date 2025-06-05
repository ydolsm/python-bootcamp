from openpyxl import load_workbook


workbook = load_workbook("sample.xlsx")
sheet = workbook["Additional"]

for row in sheet.iter_rows():
    print(row)
