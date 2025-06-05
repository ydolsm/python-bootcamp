from openpyxl import load_workbook


workbook = load_workbook("sample.xlsx")
sheet = workbook["Additional"]

for header, item in sheet.iter_rows():
    print(header.value, item.value)
