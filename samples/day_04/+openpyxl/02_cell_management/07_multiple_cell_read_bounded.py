from openpyxl import load_workbook


workbook = load_workbook("sample.xlsx")
sheet = workbook["Additional"]

for header, item in sheet.iter_rows(min_row=3, max_row=6):
    print(header.value, item.value)
