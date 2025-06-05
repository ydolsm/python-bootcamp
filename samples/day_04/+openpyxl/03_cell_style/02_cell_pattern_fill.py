from openpyxl import load_workbook
from openpyxl.styles import PatternFill

workbook = load_workbook("sample.xlsx")
sheet = workbook["Additional"]

for (cell,) in sheet["A3:A7"]:
    cell.fill = PatternFill(fill_type="solid", fgColor="4F81BD")

workbook.save("sample.xlsx")
